import os
import datetime
import openpyxl as xl
from .models.database import Table
from .settings import Settings
from .workbooks import TreeDrawer, TreeEraser


class WorkbookControl():
    """ワークブック制御"""


    def __init__(self, target, mode, settings={}, debug_write=False):
        """初期化

        Parameters
        ----------
        target : str
            ワークブック（.xlsx）へのファイルパス
        mode : str
            既存のワークブックが有ったときの挙動。 'w' は新規作成して置換え、 'a' は追記
        settings : dict
            各種設定
        """
        self._wb_file_path = target
        self._mode = mode
        self._settings_obj = Settings(dictionary=settings)
        self._debug_write = debug_write
        self._wb = None
        self._ws = None


    @property
    def workbook_file_path(self):
        return self._wb_file_path


    def render_worksheet(self, target, based_on, debug_write=False):
        """ワークシート描画

        Parameters
        ----------
        target : str
            シート名
        based_on : str
            CSVファイルパス
        debug_write : bool
            デバッグライト
        """

        if self._wb is None:
            self.ready_workbook()

        self.ready_worksheet(target=target)

        # CSV読込
        table = Table.from_csv(file_path=based_on)

        # ツリードロワーを用意、描画（都合上、要らない罫線が付いています）
        tree_drawer = TreeDrawer(table=table, ws=self._ws, settings_obj=self._settings_obj, debug_write=debug_write)
        tree_drawer.render()


        # 要らない罫線を消す
        # DEBUG_TIPS: このコードを不活性にして、必要な線は全部描かれていることを確認してください
        if True:
            tree_eraser = TreeEraser(table=table, ws=self._ws, settings_obj=self._settings_obj, debug_write=debug_write)
            tree_eraser.render()
        else:
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] eraser disabled")


    def remove_worksheet(self, target):
        """存在すれば、指定のワークシートの削除。存在しなければ無視

        Parameters
        ----------
        target : str
            シート名
        """

        if self.exists_sheet(target=target):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] remove `{target}` sheet...")

            self._wb.remove(self._wb[target])


    def save_workbook(self):
        """ワークブックの保存"""

        if self._debug_write:
            print(f"[{datetime.datetime.now()}] save `{self._wb_file_path}` file...")

        # ワークブックの保存            
        self._wb.save(self._wb_file_path)


    def ready_workbook(self):
        """ワークブックを準備する"""

        # 既存のファイルが有ったときの挙動
        if os.path.isfile(self._wb_file_path):
            # 既存のファイルへ追記
            if self._mode == 'a':
                if self._debug_write:
                    print(f"[{datetime.datetime.now()}] `{self._wb_file_path}` file exists, read.")

                self._wb = xl.load_workbook(filename=self._wb_file_path)

                return
            
            elif self._mode == 'w':
                pass
            
            else:
                raise ValueError(f"unsupported {self._mode=}")

                    
        # ワークブックを新規生成
        if self._debug_write:
            print(f"[{datetime.datetime.now()}] `{self._wb_file_path}` file not exists, create.")

        self._wb = xl.Workbook()


    def exists_sheet(self, target):
        """シートの存在確認
        
        Parameters
        ----------
        target : str
            シート名
        """
        return target in self._wb.sheetnames


    def ready_worksheet(self, target):
        """ワークシートを準備する
        
        Parameters
        ----------
        target : str
            シート名
        """

        # シートを作成
        if not self.exists_sheet(target):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] create `{target}` sheet...")

            self._wb.create_sheet(target)


        self._ws = self._wb[target]


    def to_tree_structure(self):
        """TODO 木構造モデルの作成
        
        Return
        ------
        root_node : TreeNode
            根ノード
        """
        pass


    def get_worksheet(self, sheet_name):
        """ワークシートの取得"""
        return self._wb[sheet_name]
