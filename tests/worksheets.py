import openpyxl as xl


def print_child(output_list, indent, node):
    """å†å¸°çš„ã«å­ãƒãƒ¼ãƒ‰ã‚’è¡¨ç¤º"""
    succ_indent = indent + '    '
    for child_node in node.child_nodes.values():
        # ã‚¨ãƒƒã‚¸ãƒ†ã‚­ã‚¹ãƒˆ
        if child_node.edge_text is not None:
            et = f"â”€{child_node.edge_text}â”€"
        else:
            et = 'â”€â”€'
        
        # è‘‰ãƒãƒ¼ãƒ‰
        if len(child_node.child_nodes) < 1:
            output_list.append(f"{indent}â””{et} ğŸ“„ ({child_node.leaf_th}) {child_node.text}")
        
        # ä¸­é–“ãƒãƒ¼ãƒ‰
        else:
            output_list.append(f"{indent}â””{et} ğŸ“ {child_node.text}")
            print_child(output_list=output_list, indent=succ_indent, node=child_node) # å†å¸°


class WorksheetDumpHandle():


    @staticmethod
    def nesting(name, items, sub_items):
        if 0 < len(sub_items):
            items.append(f"{name}=[{' '.join(sub_items)}]")


    @staticmethod
    def dump(worksheet, file):
        ws = worksheet

        with open(file, mode='w', encoding='utf8') as f:

            # å„åˆ—
            for column_th in range(1, ws.max_column + 1):
                column_letter = xl.utils.get_column_letter(column_th)
                items = []
                items.append(f"width={ws.column_dimensions[column_letter].width}")
                f.write(f"""\
({column_letter}) {' '.join(items)}
""")


            # å„è¡Œ
            for row_th in range(1, ws.max_row + 1):
                items = []
                items.append(f"height={ws.row_dimensions[row_th].height}")
                f.write(f"""\
({row_th}) {' '.join(items)}
""")


            # å„ã‚»ãƒ«
            for row_th in range(1, ws.max_row + 1):
                for column_th in range(1, ws.max_column + 1):
                    column_letter = xl.utils.get_column_letter(column_th)
                    cell = ws[f'{column_letter}{row_th}']

                    items = []

                    # Excel ã® UI ã®ä¸¦ã³é †
                    # --------------------
                    
                    # ãƒ•ã‚©ãƒ³ãƒˆã‚’å‡ºåŠ›
                    WorksheetDumpHandle.nesting(name='font', items=items, sub_items=WorksheetDumpHandle.parse_font(font=cell.font))

                    # ç½«ç·šã‚’å‡ºåŠ›
                    WorksheetDumpHandle.nesting(name='border', items=items, sub_items=WorksheetDumpHandle.parse_border(border=cell.border))

                    # èƒŒæ™¯è‰²ã‚’å‡ºåŠ›
                    WorksheetDumpHandle.nesting(name='fill', items=items, sub_items=WorksheetDumpHandle.parse_fill(fill=cell.fill))

                    # æ–‡å­—å¯„ã›ã‚’å‡ºåŠ›
                    WorksheetDumpHandle.nesting(name='alignment', items=items, sub_items=WorksheetDumpHandle.parse_alignment(alignment=cell.alignment))

                    f.write(f"""\
({column_letter}{row_th}) {' '.join(items)}
""")


    @staticmethod
    def parse_alignment(alignment):
        items = []
        if alignment is not None:
            if alignment.horizontal is not None:
                items.append(f"horizontal={alignment.horizontal}")

            if alignment.vertical is not None:
                items.append(f"vertical={alignment.vertical}")
        return items


    @staticmethod
    def parse_border(border):
        items = []

        if border is not None:
            if hasattr(border, 'left') and border.left is not None:
                WorksheetDumpHandle.nesting(name='left', items=items, sub_items=WorksheetDumpHandle.parse_side(side=border.left))

            if hasattr(border, 'right') and border.right is not None:
                WorksheetDumpHandle.nesting(name='right', items=items, sub_items=WorksheetDumpHandle.parse_side(side=border.right))

            if hasattr(border, 'top') and border.top is not None:
                WorksheetDumpHandle.nesting(name='top', items=items, sub_items=WorksheetDumpHandle.parse_side(side=border.top))

            if hasattr(border, 'bottom') and border.bottom is not None:
                WorksheetDumpHandle.nesting(name='bottom', items=items, sub_items=WorksheetDumpHandle.parse_side(side=border.bottom))

            if hasattr(border, 'diagonal') and border.diagonal is not None:
                WorksheetDumpHandle.nesting(name='diagonal', items=items, sub_items=WorksheetDumpHandle.parse_side(side=border.diagonal))

            if hasattr(border, 'vertical') and border.vertical is not None:
                WorksheetDumpHandle.nesting(name='vertical', items=items, sub_items=WorksheetDumpHandle.parse_side(side=border.vertical))

            if hasattr(border, 'horizontal') and border.horizontal is not None:
                WorksheetDumpHandle.nesting(name='horizontal', items=items, sub_items=WorksheetDumpHandle.parse_side(side=border.horizontal))

        return items


    @staticmethod
    def parse_color(color):
        items = []
        if color is not None:
            # rgb='00CCCCCC', indexed=None, auto=None, theme=None, tint=0.0, type='rgb'

            if color.type is not None:
                items.append(f"type={color.type}")

                if color.type == 'rgb':
                    if color.rgb is not None:
                        items.append(f"rgb={color.rgb}")

                elif color.type == 'int':
                    if color.indexed is not None:
                        items.append(f"indexed={color.indexed}")

                    if color.theme is not None:
                        items.append(f"theme={color.theme}")

                elif color.type == 'bool':
                    if color.auto is not None:
                        items.append(f"auto={color.auto}")

            if color.tint is not None:
                items.append(f"tint={color.tint}")

        return items


    @staticmethod
    def parse_fill(fill):
        items = []
        if fill is not None:
            if fill.patternType is not None:
                items.append(f"{fill.patternType}")

            # fgè‰²
            if fill.fgColor is not None:
                WorksheetDumpHandle.nesting(name='fgColor', items=items, sub_items=WorksheetDumpHandle.parse_color(color=fill.fgColor))

        return items


    @staticmethod
    def parse_font(font):
        items = []
        if font is not None:
            #name=None, charset=None, family=None, b=False, i=False, strike=None, outline=None, shadow=None, condense=None, color=<openpyxl.styles.colors.Color object>            

            if font.name is not None:
                items.append(f"name={font.name}")

            if font.charset is not None:
                items.append(f"charset={font.charset}")

            if font.family is not None:
                items.append(f"family={font.family}")

            if font.b is not None:
                items.append(f"b={font.b}")

            if font.i is not None:
                items.append(f"i={font.i}")

            if font.strike is not None:
                items.append(f"strike={font.strike}")

            if font.outline is not None:
                items.append(f"outline={font.outline}")

            if font.shadow is not None:
                items.append(f"shadow={font.shadow}")

            if font.condense is not None:
                items.append(f"condense={font.condense}")

            # è‰²
            WorksheetDumpHandle.nesting(name='color', items=items, sub_items=WorksheetDumpHandle.parse_color(color=font.color))

        return items


    @staticmethod
    def parse_side(side):
        items = []
        if side.style is not None:
            items.append(f"{side.style}")

        # è‰²
        WorksheetDumpHandle.nesting(name='color', items=items, sub_items=WorksheetDumpHandle.parse_color(color=side.color))

        return items
